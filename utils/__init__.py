from csv import excel
from os.path import sep, normpath, join
__author__ = 'HO.OPOYEN'


def log(text, stack=None):
    from kivy.app import App
    from kivy.logger import Logger
    app = App.get_running_app()
    if app:
        try:
            app.root.log(text, stack)
            print text, stack
        except AttributeError:
            Logger.warn('WARNING: No Log function found; reverting to print')
            print '\t', text, stack
    else:
        print text, stack


def alert(text, status_color=(0, 0, 0, 1), keep=False):
    from kivy.app import App
    app = App.get_running_app()
    if app:
        try:
            app.alert(text, status_color, keep)
        except AttributeError:
            print '\t', text, status_color, keep


def start_file(path):
    import os
    if hasattr(os, 'startfile'):
        os.startfile(path)
    else:
        os.system('open "%s"'%path)


def path_reader(path):
    #normalize any path to be read according to the current OS
    not_sep = '/' if sep=='\\' else '\\'
    if sep in path:
        return normpath(path)
    else:
        return normpath(path.replace(not_sep,sep))


def find_path(path):
    from os.path import isfile
    from kivy.resources import resource_find
    from conf import gamepath
    if not path:
        return path
    if path.startswith('http'):
        return path
    r = resource_find(path)
    if r:
        return r
    if not isfile(path_reader(path)):
        if isfile(join(gamepath, path_reader(path))):
            return join(gamepath, path_reader(path))
        from utils import log
        log('Source does not exist at path %s'%(path))
    else:
        return path_reader(path)


def toImage(self, bg_color=(1,1,1,0)):
    #create image widget with texture == to a snapshot of me
    from kivy.graphics import Translate, Fbo, ClearColor, ClearBuffers, Scale
    from kivy.core.image import Image as CoreImage

    if self.parent is not None:
        canvas_parent_index = self.parent.canvas.indexof(self.canvas)
        self.parent.canvas.remove(self.canvas)

    fbo = Fbo(size=self.size, with_stencilbuffer=True)

    with fbo:
        ClearColor(*bg_color)
        ClearBuffers()
        Scale(1, -1, 1)
        Translate(-self.x, -self.y - self.height, 0)

    fbo.add(self.canvas)
    fbo.draw()
    #EventLoop.idle()
    cim = CoreImage(fbo.texture, filename = '%s.png'%id(self))

    fbo.remove(self.canvas)

    if self.parent is not None:
        self.parent.canvas.insert(canvas_parent_index, self.canvas)

    return cim


def OpenXLDictReader(f):
    import openpyxl
    book  = openpyxl.reader.excel.load_workbook(f)
    sheet = book.active

    rows = sheet.max_row
    cols = sheet.max_column

    def item(i, j):
        return (sheet.cell(row=0, column=j).value, sheet.cell(row=i, column=j).value)

    return (dict(item(i,j) for j in range(cols)) for i in range(1, rows))


def XLRDDictReader(f, sheet_index=0):
    import xlrd, mmap
    book    = xlrd.open_workbook(file_contents=mmap.mmap(f.fileno(), 0, access=mmap.ACCESS_READ))
    sheet   = book.sheet_by_index(sheet_index)
    headers = dict( (i, sheet.cell_value(0, i) ) for i in range(sheet.ncols) )

    return ( dict( (headers[j], sheet.cell_value(i, j)) for j in headers ) for i in range(1, sheet.nrows) )


class Excel_Semicolon(excel):
    delimiter = ";"



#Handle color change: convert list to list & name to list


#color is a list, as it is a listproperty. It can be an rgb one or a list of mono string
def l_type(thelist): #return the common type for elt in list if uniq. else none
    types = {type(x) for x in thelist}
    if len(types) == 1:
        return types.pop()
    if types == {int,float}: #also acceptable for colors
        return float
    return None

def get_color(color, with_a=True):
    thetype = l_type(color)
    if with_a and len(color) == 4 and thetype in (int, float):
        return color
    if not(with_a) and len(color) == 3 and thetype in (int,float):
        return color
    if thetype in (str, unicode):
        color = ''.join(color)
        from webcolors import name_to_hex
        from kivy.utils import get_color_from_hex
        try:
            named = name_to_hex(color)
        except ValueError,E:
            from utils import log
            log('Unkwon color name "%s"'%color)
            return [0, 0, 0, 1] if with_a else [0, 0, 0]
        r, g, b, a = get_color_from_hex(named)
        if with_a:
            return r, g, b, a
        return r, g, b
    return [1,1,1,1]
