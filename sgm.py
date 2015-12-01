from kivy.app import App
from kivy.factory import Factory
from kivy.uix.behaviors import ButtonBehavior, ToggleButtonBehavior, FocusBehavior
from kivy.properties import BooleanProperty, ObjectProperty
from kivy.uix.slider import Slider
from kivy.uix.popup import Popup
from kivy.properties import NumericProperty, StringProperty,   DictProperty, ListProperty
from kivy.uix.treeview import  TreeViewLabel
from kivy.clock import Clock
from kivy.lang import Builder
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.treeview import TreeView
from deck import TreeViewField
from fields import BaseField
from conf import gamepath, FILE_FILTER
import os, os.path
from kivy.logger import Logger
from kivy.uix.gridlayout import GridLayout


class FolderTreeView(TreeView):
    folder = StringProperty()
    filters = ListProperty()
    rootpath = StringProperty()

    def on_rootpath(self, instance, value):
        self.load_folder(value)

    def load_folder(self, folder):
        self.folder = folder
        from os.path import join, isdir
        from os import listdir
        self.clear_widgets()
        self.root.nodes = list()

        for f in listdir(folder):
            if not isdir(join(folder,f)) and not f.endswith(tuple(self.filters)): continue
            n = self.add_node(TreeViewLabel(text=f))
            n.is_leaf = False
            n.is_loaded = False
            n.path = join(folder, f)

    def on_touch_down(self, touch):
        if self.collide_point(*touch.pos) and touch.is_double_tap and getattr(touch, 'button', -1) == "left":
            node = self.get_node_at_pos(touch.pos)
            if node == self.root:
                self.load_folder(self.folder)
        return TreeView.on_touch_down(self, touch)

    def callback(self, instance, node):
        from os import listdir
        from os.path import join, isdir
        if hasattr(node, 'path'):
            for f in listdir(node.path):
                if not isdir(join(node.path,f)):
                    continue
                n = self.add_node(TreeViewLabel(text=f), node)
                n.is_leaf = False
                n.is_loaded = False
                n.path = join(node.path, f)


    load_func = callback


class FileViewItem(ToggleButtonBehavior, BoxLayout):
    name = StringProperty()
    source = StringProperty()
    is_all_folder = StringProperty(False)
    tmplWidget = ObjectProperty()

    def on_state(self, target, state):
        self.realise()
        if state == 'down':
            if self.name.lower().endswith('.kv'):
                f = Factory.get('FileItemOptionKV')()
            else:
                f = Factory.get('FileItemOption')()
            f.is_all_folder = self.is_all_folder
            self.add_widget(f)
            from kivy.animation import Animation
            anim = Animation(size_hint_y=0.2, duration=.1)
            anim.start(f)
        else:
        #remove the choice box
            self.remove_widget(self.children[0])

    def on_press(self):
        if self.last_touch.is_double_tap:#directly add it to the pool
            if self.name.endswith('.bgp'):
                self.extract_package()
            else:
                self.add_item("1", 'normal')

    def add_item(self, qt, verso):
            from os.path import relpath
            stack = App.get_running_app().root.ids['deck'].ids['stack']
            ##########################################################
            qt = int(qt)
            if self.is_all_folder:
                #print self, self.source, self.name, self.is_all_folder
                #It is a folder, add all the imge from folder
                for name in [x for x in os.listdir(self.is_all_folder) if x.endswith(FILE_FILTER)]:
                ##for fiv in self.parent.children:
                    ##if fiv.is_all_folder: continue
                    if name.endswith(('.csv','.xlsx')):
                        continue
                    box = StackPart()
                    #box.name = fiv.name
                    box.name = name
                    box.source = os.path.join(self.is_all_folder, name)
                    box.qt = qt
                    box.verso = verso
                    if name.endswith('.kv'):
                        if self.is_all_folder.startswith(gamepath):
                            fold = relpath(self.is_all_folder, gamepath)
                        else:
                            fold = self.is_all_folder
                        box.template = "@%s"%os.path.join(fold, name)
                        box.realise()
                    stack.add_widget(box)
                    from kivy.base import EventLoop
                    EventLoop.idle()
            elif self.name.endswith('.csv'):
                App.get_running_app().root.ids.deck.load_file_csv(self.name)
            elif self.name.endswith('.xlsx'):
                App.get_running_app().root.ids.deck.load_file(self.name)
            elif self.name.endswith('.bgp'):
                self.extract_package()
            else:
                box = StackPart()
                box.name = self.name
                box.source = self.source
                box.qt = qt
                box.verso = verso
                stack.add_widget(box)
                if self.name.endswith('.kv'):
                    box.tmplWidget = self.tmplWidget
                    if self.name.startswith(gamepath):
                        fold = relpath(self.name, gamepath)
                    else:
                        fold = self.name
                    box.template = "@%s"%fold
                    box.realise()

    def realise(self,use_cache = False, *args):
        if not self.name.endswith('.kv'):
            return
        #Force the creation of an image from self.template, thourhg real display
        from kivy.clock import Clock
        #Force the creaiotn of the tmpl miniture for display
        from template import BGTemplate
        try:
            Logger.info('[SGM] Realise FileItemView calling From File')
            tmpl = BGTemplate.FromFile(self.name, use_cache)[-1]
        except IndexError:
            Logger.warn('Warning: template file %s contains no Template !!'%self.name)
            from utils import alert
            alert('Error while loading %s template'%self.name)
            return
        #App.get_running_app().root.ids['realizer'].add_widget(tmpl) #force draw of the beast
        self.tmplWidget = tmpl
        def inner(*args):
            from kivy.base import EventLoop
            EventLoop.idle()
            cim = tmpl.toImage()
            cim.texture.flip_vertical()
            self.ids['img'].texture = cim.texture
            #App.get_running_app().root.ids['realizer'].remove_widget(tmpl)
        Clock.schedule_once(inner, -1)

    def extract_package(self):
        from zipfile import ZipFile
        import os
        import tempfile
        zip_name = self.name
        zip_path = os.path.abspath(zip_name)
        temp_dir = tempfile.mkdtemp(prefix='BGM_%s'%os.path.split(self.name)[-1])
        from kivy.resources import resource_add_path
        resource_add_path(temp_dir)

        with ZipFile(zip_path, 'r') as zip_file:
            # Build a list of only the members below ROOT_PATH
            members = zip_file.namelist()
            # Extract only those members to the temp directory
            zip_file.extractall(temp_dir, members)

        from kivy.app import App
        dm = App.get_running_app().root.ids.deck

        pyfs = [m for m in members if m.endswith('.py')]
        kvfs = [m for m in members if m.endswith('.kv')]
        csvfs = [m for m in members if m.endswith('.csv')]
        xlsfs = [m for m in members if m.endswith('.xlsx')]
        #First python
        for m in pyfs:
            execfile(m)
        #then template
        for m in kvfs:
            from template import templateList
            Logger.info('[SGM] Extract Package: registering template file')
            templateList.register_file(os.path.join(temp_dir,m))
        #then deck
        for m in csvfs:
            dm.load_file_csv(os.path.join(temp_dir,m))
        for m in xlsfs:
            dm.load_file(os.path.join(temp_dir,m))
        from utils import start_file
        start_file(temp_dir)


class SpecialViewItem(FileViewItem):
    "FileView for Folder & CSV file. Usable for script ? "


class Stack(GridLayout):

    def add_part(self, widget, index=0):
        if not isinstance(widget, StackPart):
            raise TypeError('You can only add StackPart to Stack. %s received'%widget)

        GridLayout.add_widget(self, widget,index)

    def refresh(self, boxes = None, msg=""):
        if boxes is None:
            boxes = self.children[:]
        def inner(*args):
            b = boxes.pop()
            b.realise(withValue=True, use_cache=True)
            if not boxes:
                Clock.unschedule(inner)
                from utils import alert
                alert(msg)
                return False
        Clock.schedule_interval(inner, .1)


class StackPart(ButtonBehavior, BoxLayout):
    selected = BooleanProperty(False)
    row = NumericProperty(0)
    template = StringProperty()
    tmplWidget = ObjectProperty()
    name = StringProperty()
    values = DictProperty()
    source = StringProperty()
    image = ObjectProperty(False)
    layout = ObjectProperty(None)


    def realise(self,withValue = False, use_cache=False):
        #Force the creation of an image from self.template, thourhg real display
        #Skipt of computed image exists
        if self.image:
            return
        from kivy.clock import Clock
        #Force the creaiotn of the tmpl miniture for display
        from template import BGTemplate
        if not self.template:
            return
        try:
            if not use_cache: Logger.info( '[SGM]Realize StackPart calling from file')
            tmpl = BGTemplate.FromFile(self.template, use_cache)[-1]
        except IndexError:
            Logger.warn( 'Warning: template file %s contains no Template !!'%self.template)
            from utils import alert
            alert('Error while loading template %s '%self.template)
            return
        #App.get_running_app().root.ids['realizer'].add_widget(tmpl) #force draw of the beast

        if withValue:
            tmpl.apply_values(self.values)
        self.tmplWidget = tmpl

        def inner(*args):
            #Here is hould loop on the template to apply them on values
            from kivy.base import EventLoop
            EventLoop.idle()
            cim = tmpl.toImage()
            cim.texture.flip_vertical()
            self.ids['img'].texture = cim.texture
            #App.get_running_app().root.ids['realizer'].remove_widget(tmpl)
        Clock.schedule_once(inner, -1)

    def on_press(self):
        self.selected = not(self.selected)
        if self.selected: #update on dad selection
            if self.parent.last_selected and self.parent.last_selected != self:
                self.parent.last_selected.selected = False
            self.parent.last_selected = self

    def on_selected(self, instance, selected):
        if self.template:
            self.realise(True)
        if self.selected:
            from kivy.uix.boxlayout import BoxLayout

            BOX = BoxLayout(size_hint=(None, None), orientation = 'vertical', spacing=10)
            #Add Remove Button
            b = Factory.get('HiddenRemoveButton')(icon='trash')
            b.bind(on_press=lambda x: self.parent.remove_widget(self))
            #self.add_widget(b)
            from kivy.animation import Animation
            W = 90
            be = Factory.get('HiddenRemoveButton')(icon='edit')
            if self.template:#it is a template: add edit & export buttons
                def inner(*args):
                    p = Factory.get('TemplateEditPopup')()
                    p.name = self.template
                    options = p.ids['options']
                    #print 'editing options with ', self.values
                    options.values = self.values
                    options.tmplPath = self.template #trigger options building on popup
                    p.stackpart = self
                    p.open()
                    p.do_layout()
                    p.content.do_layout()
                    from kivy.clock import Clock
                    def _inner(*args):
                        prev = p.ids['preview']
                        tmpl = prev.children[0]
                        TS = tmpl.size
                        PS = prev.parent.size
                        W_ratio = float(.9*PS[0])/TS[0]
                        H_ratio = float(.9*PS[1])/TS[1]
                        ratio = min(W_ratio, H_ratio)
                        x, y = p.ids['FL'].center
                        prev.center = ratio * x, ratio * y
                        #p.ids['preview'].scale = ratio
                        #forcing x to 0
                        prev.x = 5
                        prev.center_y = y
                        from kivy.metrics import cm
                        p.tsize = options.current_selection[0].size[0]/cm(1), options.current_selection[0].size[1]/cm(1)
                        #print prev, prev.size, prev.pos
                        #print 'ratio', ratio
                        #print prev.parent, prev.parent.size, prev.parent.pos
                        #print prev.parent.parent

                    Clock.schedule_once(_inner, 0)
            else: #edit button for pure image
                def inner(*args):
                    p = Factory.get('SizeEditPopup')()
                    p.name = self.source
                    _img = self.toPILImage()
                    p.ids.width.text = str(_img.size[0])
                    p.ids.height.text = str(_img.size[1])
                    p.open()
                    def _inner(*args):
                        w,h = float(p.ids.width.text), float(p.ids.height.text)
                        if p.ids.w_metric.text == "cm":
                            from kivy.metrics import cm
                            w *= cm(1)
                        if p.ids.h_metric.text == "cm":
                            from kivy.metrics import cm
                            h *= cm(1)
                        self.setImageSize((w, h))
                    p.cb = _inner
            be.bind(on_press = inner)
            #self.add_widget(be)
            BOX.add_widget(be)
            anim = Animation(width=W, duration=.1)
            anim.start(be)
            #Img Export button
            bx = Factory.get('HiddenRemoveButton')(icon='export')
            bx.bind(on_press=lambda x: self.img_export())
            BOX.add_widget(bx)
            anim = Animation(width=W, duration=.1)
            anim.start(bx)
            BOX.add_widget(b)
            self.add_widget(BOX)
            anim = Animation(width=W, duration=.1)
            anim.start(b)
        else:
            self.remove_widget(self.children[0])
            #if self.template:
            #    self.remove_widget(self.children[0])

        self.focus = self.selected

    def Copy(self):
        blank = StackPart()
        for attr in ['template','name','values', "qt","verso", 'source', 'image']:
            setattr(blank, attr, getattr(self,attr))
        blank.ids['img'].texture = self.ids['img'].texture
        return blank

    def img_export(self, dst='export.png'):
        pim = self.toPILImage()
        pim.save(dst)
        from utils import start_file
        start_file(dst)

    def setImageSize(self, size):
        size = [int(x) for x in list(size)]
        img = self.toPILImage()
        self.setImage(img.resize(size))

    def toPILImage(self):
        item = self
        if item.image:
            return item.image
        elif item.template:
            from PIL.Image import frombuffer
            if item.tmplWidget:#it has been modified
                tmplWidget = item.tmplWidget
            else:
                from template import BGTemplate
                Logger.info( '[SGM] toPILIMage calling fromfile ')
                tmplWidget = BGTemplate.FromFile(item.template)
                if tmplWidget:
                    #only taking the last one
                    tmplWidget = tmplWidget[-1]
                else:
                    raise NameError('No such template: '+ item.template)
                print 'here to be added: adding on realizer, exporting & then removing. more tricky'
                if item.values:
                    tmplWidget.apply_values(item.values)
            cim = tmplWidget.toImage(for_print=True)
            pim = frombuffer('RGBA', cim.size, cim._texture.pixels, 'raw', 'RGBA',0,1)
        else:
            from PIL import Image
            pim = Image.open(self.source)
        return pim

    def setImage(self, pilimage):
        from PIL.Image import FLIP_TOP_BOTTOM
        from kivy.graphics.texture import Texture
        #Standard mode: flip the
        flip = pilimage.transpose(FLIP_TOP_BOTTOM)
        from img_xfos import img_modes
        ktext = Texture.create(size=flip.size)
        ktext.blit_buffer(flip.tobytes(), colorfmt=img_modes[flip.mode])
        self.ids['img'].texture = ktext
        self.image = pilimage

    def getSize(self):
        from conf import FORCE_FIT_FORMAT, card_format
        if self.image:#stack part with computed image
            return self.image.size
        elif self.tmplWidget:
            return self.tmplWidget.size
        elif self.template:
            self.realise(withValue=True, use_cache=True)
            return self.tmplWidget.size
        elif FORCE_FIT_FORMAT:
            return card_format.size
        else:
            if self.ids.img.texture:
                return self.ids.img.texture.size
            return card_format.size


class TemplateEditTree(TreeView):
    "Use in Template Edit Popup to display all possible fields"
    tmplPath = StringProperty(allownone=True)
    current_selection = ObjectProperty()
    values = DictProperty() #values from template, if any
    tmplDict = DictProperty()#link a tmplWidget to a node

    def update_tmpl(self,tmpl):
        if self.target:
            self.target.clear_widgets()
            self.target.add_widget(tmpl)
            tmpl.pos = self.target.pos
        self.current_selection = (tmpl, self.selected_node)

    def on_tmplPath(self, instance, value):
        self.current_selection = ()
        self.clear_widgets()
        self.get_root().nodes = list()
        if not value:
            return
        from template import BGTemplate
        #tmplPath is in the form [NAME][@PATH]. If path provided, load all tmpl from there. Without it, take name from library
        name, path = self.tmplPath.split('@')
        if not(name) and not(path):
            Logger.warn( 'Warning: tmpl Path is empty. stopping template edition')
        if not path:
            from template import templateList
            tmpls = [templateList[name]]
        else:
            tmpls = BGTemplate.FromFile(self.tmplPath, use_cache=True)
        for tmpl in tmpls:
            tmpl.apply_values(self.values)
            node =self.load_tmpl(tmpl)
            self.select_node(node) # will issue a template update

    def load_tmpl(self,tmpl):
        #Now add on load
        node = self.add_node(TreeViewLabel(text=tmpl.template_name, color_selected=(.6,.6,.6,.8)))
        node.is_leaf = False #add the thingy
        #point to the template
        node.template = tmpl
        self.tmplDict[tmpl] = node
        #Deal with Template Properties:
        for pname, editor in sorted(tmpl.vars.items()):
            self.add_node(TreeViewField(name=pname, editor=editor(tmpl)), node)
        #Deal with KV style elemebts
        for fname in sorted(tmpl.ids.keys()):
            if not isinstance(tmpl.ids[fname], BaseField):
                continue
            _wid = tmpl.ids[fname]
            if not _wid.editable:
                continue
            if _wid.default_attr:
                w = _wid.params[_wid.default_attr](_wid)
                if w is not None:#None when not editable
                    self.add_node(TreeViewField(pre_label=fname, name=_wid.default_attr, editor=w), node)
        self.toggle_node(node)
        return node


class TemplateEditPopup(Popup):

    def display_fields(self,do_display):
        from kivy.uix.label import Label
        for tmpl in self.ids.options.tmplDict:
            for ids in tmpl.ids:
                wid = getattr(tmpl.ids,ids)
                if do_display == 'down':
                    L = Label(text = str(ids), color = (1,0,0), font_size=30)
                    wid.add_widget(L)
                    L.center = wid.center
                    def update(*args):
                        L.center = wid.center
                    wid.bind(pos=update, size=update)
                    L.z = 100
                    from kivy.graphics import InstructionGroup, Color, Line
                    wid._df = InstructionGroup()
                    wid._df.add(Color(rgb=(1,0,0)))
                    wid._df.add(Line(rectangle =(wid.x, wid.y, wid.width, wid.height), width=4))
                    wid.canvas.after.add(wid._df)

                else:
                    wid.remove_widget(wid.children[0])
                    wid.canvas.after.remove(wid._df)

    def compute(self):
        tree = self.ids['options']
        if not tree.current_selection:
            Logger.warn( 'Warning: no template choosen. do nothing')
            return
        tmpl, node = tree.current_selection
        #Here is hould loop on the template to apply them on values
        values = tree.values
        if node:#do that only if a template has been selected. otherwise skip it
            for child_node in node.nodes:
                for child in child_node.walk(restrict=True):
                    key = getattr(child, 'target_key',None)
                    sv = getattr(child, 'stored_value',None)
                    print key, sv, tmpl.vars
                    if key is not None and sv is not None: # means somthing has changed
                        if child.target_attr in tmpl.vars: #just a tmpl variable
                            values[key] = sv
                        else:
                             #values["%s.%s"%(child.target_attr, key)] = sv
                        #by definitoin, key is default_attr
                            values[child.target_attr] = sv
        self.stackpart.values = values
        self.stackpart.tmplWidget = tmpl
        oldname = self.stackpart.template
        if '@' in oldname:
            oname,opath = oldname.split('@')
        else:
            oname,opath = oldname,""
        if opath:
            self.stackpart.template = '%s@%s'%(tmpl.template_name, opath)
        else:
            self.stackpart.template = tmpl.template_name
        cim =  tmpl.toImage()
        cim.texture.flip_vertical()
        self.stackpart.ids['img'].texture = cim.texture
        self.stackpart.image = False


class BGDeckMaker(BoxLayout):
    cancel_action = BooleanProperty(False)

    tmplsLib = ObjectProperty()

    def empty_stack(self):
        self.ids['stack'].clear_widgets()
        self.ids.stack.last_selected = False
        self.record_last_file("")

    def prepare_print(self, dst):
        "Launch PDF preparation. First determine the best method for placing images"

        from conf import CP
        from utils import alert
        if CP.getboolean('Print','AUTOCSV'):
            alert('Auto saving XLS deck')
            self.export_file(dst.replace('.pdf','.xlsx'))
        from printer import prepare_pdf
        from conf import card_format
        FFORMAT = (card_format.width, card_format.height)
        #3 possibilites
        # If FORCE_FOMART, just use it with auto layout placement
        if self.ids['force_format'].active:
            mode = 'FORCED'
        else:
            #Loop on all stack & check their size.
            #IF ALL have layout, launch print with layout
            #ELSE
            USE_LAYOUT = True
            sizes = set()
            WARNING = False
            for cs in self.ids.stack.children:
                if cs.layout:
                    WARNING = True
                else:
                    USE_LAYOUT = False
                    if WARNING:
                        from utils import alert
                        alert('Can not have both with and without layout (%s was without)!'%cs)
                        return
                if cs.template:
                    if cs.tmplWidget:
                        sizes.add(tuple(cs.tmplWidget.size))
                    else:
                        from template import BGTemplate
                        sizes.add(tuple(BGTemplate.FromFile(cs.template, use_cache=True)[-1].size))
                elif cs.image:
                    sizes.add(self.image.size)
                else:
                    sizes.add(FFORMAT)
            if USE_LAYOUT:
                mode = 'LAYOUT'
            else:
                if len(sizes) == 1:
                    mode = sizes.pop()
                else:
                    mode = 'BINCAN'
        #Now add the advancement gauge
        progress = self.ids['load_progress']
        progress.value = 0
        #ensure the stop button is reachable
        self.ids['stop_action'].width = 80
        self.ids['stop_action'].text = 'Stop'
        self.cancel_action = False
        size, book = prepare_pdf(stack=self.ids['stack'], dst=dst, console_mode= False, mode = mode)
        progress.max = size
        from kivy.clock import Clock
        step_counter = range(size)

        #Now ensure that ALL stackpart from front & back are realized, while re-creating cache
        from template import templateList
        templateList.templates = dict()
        SS = book.stack[0] + book.stack[1]

        def inner(*args):
            step_counter.pop()
            progress.value += 1
            #print 'remaninig index', len(book.index)
            book.generation_step()
            if (not step_counter) or self.cancel_action:
                self.cancel_action = False
                Clock.unschedule(inner)
                self.ids['stop_action'].width = 0
                self.ids['stop_action'].text = ''
                progress.value = 0
                book.save()
                book.show()
                from utils import alert
                alert('PDF Export completed')
                return False
            else:
                Clock.schedule_once(inner,0.01)

        def realize_inner(*args):
            s = SS.pop()
            s.realise(withValue=True, use_cache=True)
            if SS:
                Clock.schedule_once(realize_inner, 0.01)
            else:
                Clock.schedule_once(inner,.1)

        Clock.schedule_once(inner,.1)
        return True

    def load_template_lib(self, force_reload = False, background_mode = False):
        #Same as load_folder('/templates') but with delay to avoid clash
        progress = self.ids['load_progress']
        pictures = self.ids['pictures']
        if background_mode:
            pictures = Factory.get('PictureGrid')()
        pictures.clear_widgets()
        if self.tmplsLib and not force_reload:
            for c in reversed(self.tmplsLib):
                if c.parent:
                    c.parent.remove_widget(c)
                pictures.add_widget(c)
            return
        from template import templateList
        tmpls = sorted([x for x in os.listdir('templates') if x.endswith('.kv')], key=lambda x:x.lower() , reverse=True)
        C = len(tmpls)
        progress.max = C
        progress.value = 1

        def inner(*args):
            if not tmpls:
                Clock.unschedule(inner)
                self.tmplsLib = pictures.children[:]
                return
            _f = os.path.join('templates',tmpls.pop())
            templateList.register_file(_f)
            img = FileViewItem(source="", name=_f)
            pictures.add_widget(img)
            progress.value += 1
            img.realise()
        Clock.schedule_interval(inner, .1)

    def load_folder(self, folder):
        if not folder:
            return
        progress = self.ids['load_progress']
        pictures = self.ids['pictures']
        pictures.clear_widgets()
        C = len([x for x in os.listdir(folder) if x.endswith(FILE_FILTER)])
        progress.max = C
        progress.value = 1
        #pg = Factory.get('PictureGrid')()
        pictures.add_widget(SpecialViewItem(source='folder-add', is_all_folder=folder, name="Add %d Items"%C))
        docs = sorted([x.lower() for x in os.listdir(folder) if x.endswith(FILE_FILTER)], reverse=True)
        #ensure the stop button is reachable
        self.ids['stop_action'].width = 80
        self.ids['stop_action'].text = 'Stop'
        self.cancel_action = False
        def inner(*args):
            if docs:
                _f = docs.pop()
                if _f.endswith('.kv'): #it is a template:
                    source = 'img/card_template.png'
                    _f = os.path.join(folder, _f)
                elif _f.endswith(('.csv',".xlsx",".bgp")):
                    _f = os.path.join(folder, _f)
                else:
                    source = os.path.join(folder,_f)
                if _f.endswith(('.csv','.xlsx')):
                    img = SpecialViewItem(source= 'csv', name=_f)
                elif _f.endswith('.bgp'):
                    img = SpecialViewItem(source = 'cubes', name = _f)
                else:
                    img = FileViewItem(source=source, name=_f)
                pictures.add_widget(img)
                progress.value += 1
                if _f.endswith('.kv'):
                    img.realise()
            if self.cancel_action or not docs:
                self.cancel_action = False
                Clock.unschedule(inner)
                self.ids['stop_action'].width = 0
                self.ids['stop_action'].text = ''
                progress.value = 0
                return False
        Clock.schedule_interval(inner,.025)

    def choose_file_popup(self,title, cb):
        from kivy.uix.popup import Popup
        from kivy.uix.filechooser import FileChooserListView
        from conf import get_last_dir, set_last_dir
        f = FileChooserListView(path =get_last_dir(self))
        def valid(*args):
            p.dismiss()
        f.bind(on_submit= valid)
        p = Popup(content=f)
        p.title = title
        def inner(*args):
            if p.content.selection:
                selection = p.content.selection[0]
                from os.path import split
                set_last_dir(self,split(selection)[0])
                cb(selection)
        p.on_dismiss = inner
        p.open()

    def write_file_popup(self,title,cb, default='export.pdf'):
        from conf import CP
        lf = CP.get('Path','last_file')
        if default.endswith(('.csv','.xlsx')): #try to use last file if exsits
            default = lf or default
        elif default.endswith('.pdf'):
            if lf:
                default = lf.replace('.xlsx','.pdf').replace('.csv','.pdf')
        p = Factory.get('WriteFilePopup')()
        p.title = title
        p.cb = cb
        p.default_name = default
        p.open()

    def record_last_file(self,filepath):
        #set path as last_opoend_file
        from conf import CP, gamepath
        from os.path import relpath
        if filepath.startswith(gamepath):
            fpath = relpath(filepath,gamepath)
        else:
            fpath = filepath
        CP.set('Path','last_file',fpath)
        CP.write()

    def load_file(self, filepath = 'myxlsfile.xlsx'):
        stack = self.ids['stack']
        from kivy.resources import resource_add_path
        from os.path import split
        resource_add_path(split(filepath)[0])
        from utils import XLRDDictReader
        from utils import find_path
        self.record_last_file(filepath)
        boxes = list()
        with open(filepath, 'rb') as XLFile:
            stds_headers = {'qt', 'source', 'template', 'dual', 'layout_x', 'layout_y',
                            'layout_w', 'layout_h', 'layout_angle', 'layout_page'
            }
            for obj in XLRDDictReader(XLFile):
                box = StackPart()
                box.qt = int(obj.get('qt', 0))
                box.dual = 'dual' in obj and obj['dual']
                values = dict()
                if "template" in obj and obj['template']:
                    box.template = obj['template']
                if 'source' in obj and obj['source']:
                    _s = find_path(obj['source'])
                    if _s is None:
                        _s = obj['source']
                    box.source = _s
                if 'layout_x' in obj and 'layout_y' in obj and 'layout_w' in obj and 'layout_h' in obj and 'layout_angle' in obj and 'layout_page' in obj:
                    box.layout = obj['layout_x'], obj['layout_y'], obj['layout_w'], obj['layout_h'], obj['layout_angle'], obj['layout_page']
                for attr in obj:
                    if attr in stds_headers:
                        continue
                    if obj[attr]:
                        values[attr] = obj[attr]
                box.values = values
                stack.add_widget(box)
                boxes.append(box)
            boxes.reverse()

            self.refresh(boxes, msg='Import %s over'%split(filepath)[-1])

    def refresh(self, boxes=None, msg=""):
        if boxes is None:
            boxes = self.ids.stack.children[:]
        progress = self.ids['load_progress']
        progress.value = 0
        progress.max = len(boxes)
        #ensure the stop button is reachable
        self.ids['stop_action'].width = 80
        self.ids['stop_action'].text = 'Stop'
        self.cancel_action = False

        def inner(*args):
            progress.value += 1
            b= boxes.pop()
            b.realise(withValue=True, use_cache=True)
            if self.cancel_action or not boxes:
                self.cancel_action = False
                Clock.unschedule(inner)
                self.ids['stop_action'].width = 0
                self.ids['stop_action'].text = ''
                progress.value = 0
                from utils import alert
                from os.path import split
                alert(msg)
                return False

        Clock.schedule_interval(inner, .1)

    def load_file_csv(self, filepath='mycsvfile.csv'):
        #Now also CSV export
        import csv
        from utils import find_path
        self.record_last_file(filepath)
        with open(filepath, 'rb') as csvfile:
            dialect = csv.Sniffer().sniff(csvfile.read(1024), delimiters=";,")
            csvfile.seek(0)
            reader = csv.DictReader(csvfile, dialect=dialect)
            header = reader.fieldnames
            stack = self.ids['stack']
            remaining_header = set(header) - {'qt', 'source', 'template', 'dual', 'layout'}
            boxes = list()
            for index, obj in enumerate(reader):
                #print index, obj
                if set([obj[_v] for _v in obj]) == set(dialect.delimiter):#skipping empty row made of ;;;;;; or ,,,,,
                    continue
                qt = 1
                if 'qt' in obj:
                    qt = int(obj['qt'])
                verso = 'normal'
                if 'dual' in obj:
                    if not obj['dual'] or obj['dual'].lower() == 'false':
                        verso = 'normal'
                    else:
                        verso = 'down'
                box = StackPart()
                box.verso = verso
                box.qt = qt
                if 'template' in obj and obj['template']:
                    box.template = obj['template']
                if 'source' in obj and obj['source']:
                    box.source = find_path(obj['source'])
                if 'layout' in obj and obj['layout']:
                    box.layout = obj['layout']
                values = dict()
                for attr in remaining_header:
                    v = obj.get(attr, '')
                    if v is not '':#'' means cell is empty
                        if isinstance(v, basestring):
                            v= unicode(v, 'latin-1')
                        values[attr] = v
                box.values = values
                stack.add_widget(box)
                boxes.append(box)
            boxes.reverse()

            progress = self.ids['load_progress']
            progress.value = 0
            progress.max = len(boxes)
            #ensure the stop button is reachable
            self.ids['stop_action'].width = 80
            self.ids['stop_action'].text = 'Stop'
            self.cancel_action = False

            def inner(*args):
                progress.value +=1
                b= boxes.pop()
                b.realise(withValue=True)
                if self.cancel_action or not boxes:
                    self.cancel_action = False
                    Clock.unschedule(inner)
                    self.ids['stop_action'].width = 0
                    self.ids['stop_action'].text = ''
                    progress.value = 0
                    from utils import alert
                    from os.path import split
                    alert('Import %s over'%split(filepath)[-1])
                    return False

            Clock.schedule_interval(inner, .1)

    def export_file_CSV(self, cards, filepath, HAS_LAYOUT):
        #Now also CSV export
        import csv
        if cards:
            my_dict = {}
            #update the dict with all possible 'values' keys
            for c in cards:
                if c['values']:
                    my_dict.update(**c['values'])
            with open(filepath, 'wb') as f:  # Just use 'w' mode in 3.x
                fields_order = ['qt','dual','source','template'] + my_dict.keys()[:]
                if HAS_LAYOUT:
                    fields_order = ['qt','dual','source', 'layout', 'template'] + my_dict.keys()[:]
                from utils import Excel_Semicolon
                w = csv.DictWriter(f, fields_order, dialect=Excel_Semicolon)
                w.writeheader()
                for c in cards:
                    for k in my_dict:
                        my_dict[k]=None
                    my_dict.update(c)
                    del my_dict['values']
                    my_dict.update(**c['values'])
                    w.writerow(my_dict)

    def export_file(self, filepath='myxlfile.xlsx'):
        from collections import OrderedDict
        from conf import gamepath
        from os.path import relpath, isfile, splitext, split
        self.record_last_file(filepath)
        FOLDER, FNAME = split(filepath)
        FNAME = splitext(FNAME)[0]
        od = OrderedDict()
        cards = list()
        HAS_LAYOUT = False
        for item in reversed(self.ids['stack'].children):
            if not isinstance(item, Factory.get('StackPart')): continue
            d = dict()
            d['qt'] = item.qt
            if item.source:
                if isfile(item.source): #it is a full path. convert it
                    _s = relpath(item.source, gamepath)
                    if item.source.startswith(FOLDER):
                        _s = relpath(item.source, FOLDER)
                else:
                    _s = item.source
                if item.template and item.source == 'img/card_template.png':
                    _s = ""
                d['source'] = _s
            else:
                d['source'] = ""
            if item.template:
                d['template'] = item.template
            d['dual'] = not(item.verso == 'normal')
            d['values'] = item.values
            if item.layout:
                HAS_LAYOUT = True
                #d['layout'] = item.layout
                _x,_y,_w,_h,_a, _p = item.layout
                d['layout_w'] = _w
                d['layout_h'] = _h
                d['layout_y'] = _y
                d['layout_x'] = _x
                d['layout_page'] = _p
                d['layout_angle'] = _a

            cards.append(d)
        od['cards'] = cards
        from conf import CP
        mode = CP.get('Export','format')
        if mode =='csv':
            self.export_file_CSV(cards, filepath, HAS_LAYOUT)
            return
        elif mode == 'xlsx':
            #Now also XL export
            import openpyxl
            from openpyxl.utils import get_column_letter
            if cards:
                my_dict = {}
                #update the dict with all possible 'values' keys
                for c in cards:
                    if c['values']:
                        my_dict.update(**c['values'])
                wb= openpyxl.Workbook()
                ws = wb.active
                ws.title = FNAME
                fields_order = ['qt','dual','source','template'] + sorted(my_dict.keys())
                if HAS_LAYOUT:
                    fields_order = ['qt','dual','source', 'layout_x','layout_y','layout_w','layout_h', 'layout_angle','layout_page', 'template'] + sorted(my_dict.keys())
                for colindex, h in enumerate(fields_order):
                    ws['%s%s'%(get_column_letter(colindex+1),1)] = h
                for rowindex,c in enumerate(cards):
                    for k in my_dict:
                        my_dict[k]=None
                    my_dict.update(c)
                    del my_dict['values']
                    my_dict.update(**c['values'])
                    for colindex,h in enumerate(fields_order):
                        _v = my_dict[h]
                        ws['%s%s' % (get_column_letter(colindex+1),rowindex+2)] = _v
            wb.save(filepath)
        from utils import alert
        from os.path import split
        alert('Exporting %s over'%split(filepath)[-1])

    def compute_stats(self,grid):
        if grid is None:
            grid = self.ids['stack']
        qt = 0
        qt_front = 0
        qt_back = 0
        for index,c in enumerate(grid.children):
            if not isinstance(c, Factory.get('StackPart')):
                continue
            c.row = len(grid.children)-index
            qt += c.qt
            if c.verso == 'normal':
                qt_front += c.qt
            else:
                qt_back += c.qt
        num_part = len([_c for _c in grid.children if isinstance(_c, StackPart)])
        label = "Stack made of %s parts / %s Cards: %s Front - %s Back"%(num_part,qt,qt_front,qt_back)
        self.ids['stats'].text = label

    def linearize(self,progressbar = None):
        #transform all stack with qt>1 to as many stack with qt 1
        cs = self.ids.stack.children[:]
        cs.reverse()
        dst = list()
        if progressbar:
            from kivy.base import EventLoop
            progressbar.value = 0
            progressbar.max = len(cs) * 2
        for c in cs:
            if progressbar:
                progressbar.value+=1
                EventLoop.idle()
            for _ in range(c.qt):
                d = c.Copy()
                d.qt = 1
                dst.append(d)
        self.ids.stack.clear_widgets()
        for d in dst:
            if progressbar:
                progressbar.value += 1
                EventLoop.idle()
            self.ids.stack.add_widget(d)
            d.realise(use_cache=True)


Builder.load_file('kv/sgm.kv')

class SGMApp(App):
    def compute_stats(self,grid): return self.root.compute_stats(grid)
if __name__ == '__main__':
    SGMApp().run()
